from openpyxl import load_workbook
import optparse

parser = optparse.OptionParser('usage: -s -e -v')
parser.add_option('-s', dest='colIniz', type='string', help='colonna iniziale')
parser.add_option('-e', dest='colFin', type='string', help='colonna finale')
parser.add_option('-r', dest='rowStart', type='string', help='riga iniziale')
(options, args) = parser.parse_args()
if options.colIniz==None or options.colFin==None or options.rowStart==None:
    print(parser.usage)
    exit(0)

wb = load_workbook(filename = 'prova_biblioteca.xlsx')
ws = wb.active

colIniz = options.colIniz
colFin = options.colFin

row = int(options.rowStart)
print('[')
while ws[colIniz+str(row)].value != None:
	print('{')
	print('"inventario": "' + str(ws['A'+str(row)].value) + '",')
	print('"autore": "' + str(ws['B'+str(row)].value) + '",')
	print('"titolo": "' + str(ws['C'+str(row)].value) + '",')
	print('"citta": "' + str(ws['D'+str(row)].value) + '",')
	print('"editore": "' + str(ws['E'+str(row)].value) + '",')
	print('"data": "' + str(ws['F'+str(row)].value) + '",')
	print('"volume": "' + str(ws['G'+str(row)].value) + '",')
	print('"collocazione": "' + str(ws['H'+str(row)].value) + '",')
	print('"provenienza": "' + str(ws['I'+str(row)].value) + '",')
	print('"valore": "' + str(ws['J'+str(row)].value) + '",')
	print('"data ingresso": "' + str(ws['K'+str(row)].value) + '",')
	print('"vecchi dati": "' + str(ws['L'+str(row)].value) + '",')
	print('"note": "' + str(ws['M'+str(row)].value) + '"' )
	print('},')
	row += 1
print(']')
